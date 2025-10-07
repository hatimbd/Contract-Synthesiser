import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border, Alignment, NamedStyle
import copy

def get_next_version(filepath):
    """Détermine le nom de la prochaine version (V1, V2, ...)."""
    if not os.path.exists(filepath):
        return "V1"
    wb = load_workbook(filepath)
    versions = [int(s[1:]) for s in wb.sheetnames if s.startswith("V") and s[1:].isdigit()]
    wb.close()
    return f"V{max(versions)+1}" if versions else "V1"

def copy_sheet_format(source_sheet, target_sheet):
    """
    Copie la mise en forme (largeurs de colonnes, styles de cellules) de source_sheet
    vers target_sheet *sans* recopier les valeurs.
    """
    from openpyxl.utils import get_column_letter

    # Copier les largeurs de colonnes (avec vérification)
    max_col = source_sheet.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        col_dim = source_sheet.column_dimensions.get(col_letter)
        if col_dim is not None and getattr(col_dim, "width", None):
            target_sheet.column_dimensions[col_letter].width = col_dim.width

    # Copier les styles cellule par cellule
    for row in source_sheet.iter_rows():
        for cell in row:
            tgt = target_sheet.cell(row=cell.row, column=cell.column)
            try:
                if cell.has_style:
                    tgt._style = cell._style
            except Exception:
                pass

def update_excel(changes, output_path):
    import copy
    from openpyxl.styles import PatternFill

    version = get_next_version(output_path)
    temp_path = output_path.replace(".xlsx", "_temp.xlsx")

    if not os.path.exists(output_path):
        print(" Le fichier Excel n'existe pas. Veuillez le placer dans le dossier output.")
        return

    wb = load_workbook(output_path)
    candidate_sheets = [s for s in wb.sheetnames if s.startswith("V") and s[1:].isdigit()]
    last_version = sorted(candidate_sheets, key=lambda x: int(x[1:]))[-1] if candidate_sheets else wb.sheetnames[-1]

    source_sheet = wb[last_version]
    new_sheet = wb.create_sheet(title=version)
    copy_sheet_format(source_sheet, new_sheet)

    # Lire les données de la feuille source
    data = [
        list(row)
        for row in source_sheet.iter_rows(values_only=True)
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]

    if not data:
        print(" La feuille source est vide.")
        return

    headers = [str(h).strip() for h in data[0]]
    df = pd.DataFrame(data[1:], columns=headers)

    if "Log_ID" not in df.columns:
        print(" Aucune colonne Log_ID trouvée dans le fichier Excel.")
        return

    # --- Styles de surlignage ---
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")   # rouge clair
    pink_fill = PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid")  # rose clair

    # --- Appliquer les changements cellule par cellule ---
    for change in changes:
        log_id = str(change.get("Log_ID", "")).strip()
        col = change.get("column")
        new_value = change.get("new_value", "")
        action = change.get("action")

        if not log_id or col not in df.columns:
            continue

        if action == "ADD_UPDATE":
            # Modifier ou créer la ligne
            if log_id in df["Log_ID"].astype(str).values:
                df.loc[df["Log_ID"].astype(str) == log_id, col] = new_value
            else:
                new_row = {c: "" for c in df.columns}
                new_row["Log_ID"] = log_id
                new_row[col] = new_value
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        elif action == "DELETE":
            # Vider la cellule (ne pas supprimer la ligne)
            if log_id in df["Log_ID"].astype(str).values:
                df.loc[df["Log_ID"].astype(str) == log_id, col] = ""

    # --- Écriture dans la nouvelle feuille ---
    for c_idx, h in enumerate(df.columns.tolist(), start=1):
        new_sheet.cell(row=1, column=c_idx, value=h)

    # Ligne modèle (style)
    model_row_idx = 2 if source_sheet.max_row >= 2 else 1
    model_row = list(source_sheet.iter_rows(min_row=model_row_idx, max_row=model_row_idx))[0]
    model_styles = [copy.copy(cell._style) for cell in model_row]

    # --- Écriture des données avec surlignage contextuel ---
    for r_idx, row_vals in enumerate(df.values.tolist(), start=2):
        for c_idx, value in enumerate(row_vals, start=1):
            cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
            if c_idx <= len(model_styles):
                cell._style = copy.copy(model_styles[c_idx - 1])

            log_id = str(df.iloc[r_idx - 2]["Log_ID"]).strip()
            col_name = df.columns[c_idx - 1]

            # Identifier le type de modification pour cette cellule
            for ch in changes:
                if ch["Log_ID"] == log_id and ch["column"] == col_name:
                    if ch["action"] == "ADD_UPDATE":
                        cell.fill = red_fill
                    elif ch["action"] == "DELETE":
                        cell.fill = pink_fill
                    break  # optimisation

    # --- Sauvegarde ---
    wb.save(temp_path)
    wb.close()

    try:
        if os.path.exists(output_path):
            os.remove(output_path)
        shutil.move(temp_path, output_path)
        print(f" Nouvelle version créée : {version}")
    except PermissionError:
        alt = output_path.replace(".xlsx", f"_{version}_new.xlsx")
        shutil.move(temp_path, alt)
        print(f" Fichier verrouillé. Sauvegarde sous : {alt}")
